# _*_ coding:utf-8 _*_
import sys
import time
import urllib
import urllib2
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook

reload(sys)
sys.setdefaultencoding('utf8')
#some User Agents
hds = [{'User_Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.110 Safari/537.36'},{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'}]

def book_spider(book_tag):
    page_num = 0
    book_list = []
    try_times = 0
    f = open('lxm','w')

    while(1):
        url = 'https://www.douban.com/tag/'+urllib.quote(book_tag)+'/book?start='+str(page_num*20)
        #url = 'https://www.douban.com/tag/'+urllib.quote(book_tag)+'?start='+str(page_num*20)+'&type=T'
        time.sleep(np.random.rand()*5)

        try:
            req = urllib2.Request(url, headers = hds[page_num%len(hds)])
            source_code = urllib2.urlopen(req).read()
            plain_text = str(source_code)
            f.write(plain_text)

        except (urllib2.HTTPError, urllib2.URLError), e:
            print e
            continue
        
        soup = BeautifulSoup(plain_text)
        list_soup = soup.find('div', {'class':'mod book-list'})

        try_times += 1
        if list_soup == None and try_times < 200:
            continue
        elif list_soup == None or len(list_soup)<=1:
            break#200次请求后没有新的内容

        for book_info in list_soup.findAll('dd'):
            title = book_info.find('a', {'class':'title'}).string.strip()
            desc = book_info.find('div', {'class':'desc'}).string.strip()
            desc_list = desc.split('/')
            book_url = book_info.find('a', {'class':'title'}).get('href')
            print book_url
            try:
                author_info = '作者/译者：'+'/'.join(desc_list[0:-3])
            except:
                author_info = '作者/译者：暂无'

            try:
                pub_info = '出版信息：'+'/'.join(desc_list[-3:])

            except:
                pub_info = '出版信息：暂无'

            try:
                rating = book_info.find('span',{'class':'rating_nums'}).string.strip()

            except:
                rating='0.0'

            try:
                people_num = get_people_num(book_url)
                people_num = people_num.strip('人评价')

            except:
                people_num = '0'

            book_list.append([title,rating, people_num, author_info, pub_info])
            try_times=0#set 0 when got valid information
        page_num+=1
        print 'Downloading Information From Page %d' %page_num
    return book_list
def read_read(url):
    try:
        req = urllib2.Request(url, headers=hds[np.random.randint(0, len(hds))])
        print req
        source_code = urllib2.urlopen(req).read()
        print source_code

    except BaseException,b:
        print b
def get_people_num(url):
    try:
        req = urllib2.Request(url, headers=hds[np.random.randint(0, len(hds))])
        source_code = urllib2.urlopen(req).read()
        plain_text = str(source_code)
        print 'plain_text'
    except(urllib2.HTTPError, urllib2.URLError), e:
        print e
    except BaseException,b:
        print 'baseexception',b
    print 'the soup-----------1'
    soup = BeautifulSoup(plain_text)
    print 'the soup-----------2'
    
    people_num = soup.find('div',{'class':'rating_sum'}).findAll('span')[1].string.strip()
    print 'the soup-----------3'
    print people_num,'people_num'
    return people_num

def do_spider(book_tag_lists):
    book_lists = []
    for book_tag in book_tag_lists:
        book_list = book_spider(book_tag)
        book_list = sorted(book_list, key=lambda x: x[1], reverse=True)
        book_lists.append(book_list)
    return book_lists

def print_book_lists_excel(book_lists, book_tag_lists):
    wb = Workbook()
    ws=[]
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i].decode()))

    for i in range(len(book_tag_lists)):
        ws[i].append(['序号','书名','评分', '评价人数','作者','出版社'])
        count=1
        for b1 in book_lists[i]:
            ws[i].append([count, b1[0], float(b1[1]), int(b1[2]),b1[3],b1[4]])
            count+=1
    save_path = 'book_list'

    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i].decode())
    save_path+='.xlsx'
    wb.save(save_path)

if __name__ == '__main__':
    #book_tag_lists = ['科普','经典','生活','心灵','文学']
    book_tag_lists = ['摄影']
    book_lists = do_spider(book_tag_lists)
    print_book_lists_excel(book_lists, book_tag_lists)
        
