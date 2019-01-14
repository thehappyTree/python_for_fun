# -*- coding:utf-8 -*-
import json
from openpyxl import load_workbook,Workbook

class Exchange:
    def __init__(self,parent_path,item,product,taocard,recharge):
        self.parent_path = parent_path
        #p_tao = self.parent_path +'/' +'cardBrand_courseCard_taosubjects.json'
        #p_sub = self.parent_path +'/' + 'cardBrand_courseCard_subjects.json'
        #p_rec = self.parent_path +'/' + 'card_recharge.json'
        #p_goods = self.parent_path +'/' + 'goods_goodsType_listCombineGoods.json'
        p_tao = self.parent_path +'/' + taocard
        p_sub = self.parent_path +'/' + item
        p_rec = self.parent_path +'/' + recharge
        p_goods = self.parent_path +'/' + product
        self._tao = json.load(open(p_tao,'r'))
        self._sub = json.load(open(p_sub,'r'))
        self._rec = json.load(open(p_rec,'r'))
        self._goods = json.load(open(p_goods,'r'))

    def getExcelSheet(self,index):
        wb = load_workbook('MerchantDataImportTemplateV1.0.xlsx')
        sheet_names = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheet_names[index])# index为0为第一张表 
        return ws

    def action(self):
        wb = Workbook()
        ws = wb.active
        #ws = self.getExcelSheet(1)
        ws.append([u'编号',u'大类',u'小类'])
        content = []
        #print self._sub
        bodys = self._sub['body']
        # bodys为大类级别
        start = 1001
        for body in bodys:
            classCl = {}
            classCl['categoryCode'] = body['categoryCode']#大类code
            classCl['bigClass']= body['categoryName']#大类名字
            # 每个大类下面的小类以及小类下的具体项目
            items = body['subjects']
            smallItem = set()
            # items为小类级别
            for item in items:
                # 小类名称       
                smallItem.add((item['objName'],item['code']))

            classCl['smallItem'] = smallItem
            content.append(classCl)

        bodys_good = self._goods['body']
        goodContent = []
        for good in bodys_good:
            goodCl = {}
            goodCl['goodsTypeName'] = good['goodsTypeName']
            goodCl['goodsTypeCode'] = good['goodsTypeCode']
            if good['subjects']:
                goodSet = set()
                for s in good['subjects']:
                    goodSet.add((s['objName'],s['code']))
                goodCl['objName'] = goodSet
            else:
                continue
            goodContent.append(goodCl)

        #taosub
        bodys_tao = self._tao['body']
        taoContent = []
        for tao in bodys_tao:
            taoCl = {}
            if tao['taoSubList']:
                taoSet = set()
                for s in tao['taoSubList']:
                    taoCl['categoryName'] = s['categoryName']
                    taoCl['objName'] = s['objName']
                    taoCl['code'] = s['code']
                    #goodCl['objName'] = goodSet
                    taoContent.append(taoCl)

        #rec
        bodys_rec = self._rec['body']
        recContent = []
        for rec in bodys_rec:
            recCl = {}
            recCl['singleRuleValue'] = rec['singleRuleValue']
            memb = rec['memberCard']
            recCl['objName'] = memb['objName']
            recCl['code'] = memb['code']
            recContent.append(recCl)

        count = 1001
        excList = [None, None, None, None,None] 
        for c in content:
            excList[0] = count
            excList[1] = c['bigClass']
            for s in c['smallItem']:
                excList[3] = s[1]
                excList[2] = s[0]
                count += 1
                ws.append(excList)
        for tao in taoContent:
            excList[0] = count
            excList[1] = tao['categoryName']
            excList[2] = tao['objName']
            excList[3] = tao['code']
            ws.append(excList)

        for g in goodContent:
            excList[0] = count
            excList[1] = g['goodsTypeName']
            for s in g['objName']:
                excList[3] = s[1]
                excList[2] = s[0]
                count += 1
                ws.append(excList)

        for r in recContent:
            excList[0] = count
            excList[1] = r['singleRuleValue']
            excList[2] = r['objName']
            excList[3] = r['code']
            ws.append(excList)
        self.kind5(self._sub['body'],self._tao['body'])
        self.kind7(self._goods['body'])
        self.kind9(self._rec['body'])
        wb.save(self.parent_path+'/feilei.xlsx')

    def kind5(self,content_sub,content_tao):
        wb = Workbook()
        ws = wb.active
        ws.append([u'编号',u'大类',u'小类',u'价格',u'耗时'])
        #cCls = [None,None,None,None,None,None]
        count = 1
        for c in content_sub:
            sub = c['subjects']
            for s in sub:
                cCls = []
                cCls.append(count)
                #cCls[1] = s['code']
                cCls.append(s['categoryName'])
                cCls.append(s['objName'])
                cCls.append(int(s['price'])/100.00)
                cCls.append(s['times'])
                ws.append(cCls)
                count+=1

        for c in content_tao:
            tSub = c['taoSubList']
            for t in tSub:
                cCls = []
                cCls.append(count)
                #cCls.append(t['code'])
                cCls.append(t['categoryName'])
                cCls.append(t['objName'])
                cCls.append(int(t['price'])/100.00)
                cCls.append(t['times'])
                ws.append(cCls)
                count+=1
        wb.save(self.parent_path+'/xiangmu.xlsx')
        
    def kind7(self,content_goods):
        wb = Workbook()
        ws = wb.active
        cCls = [None,None,None,None,None,None,None]
        ws.append([u'编号',u'大类',u'子类',u'存货量',u'单位',u'价格'])
        count = 1
        for c in content_goods:
            if c['subjects']:
                for s in c['subjects']:
                    cCls = []
                    cCls.append(count)
                    cCls.append(s['goodsTypeName'])
                    cCls.append(s['objName'])
                    #cCls.append(s['code'])
                    cCls.append(s['stockNum'])
                    cCls.append(s['goodsUnit'])
                    cCls.append(int(s['price'])/100.00)
                    ws.append(cCls)
        wb.save(self.parent_path + '/goods.xlsx')

     
    def kind9(self,cont_rec):
        wb = Workbook()
        ws = wb.active
        ws.append([u'编号',u'分类',u'卡名称',u'价格',u'开始时间',u'结束时间'])
        #cCls = [None,None,None,None,None,None,None]
        count = 2001
        for c in cont_rec:
            if c['memberCard']:
                cCls = []
                name = c['singleRuleValue']
                cCls.append(count)
                mem = c['memberCard']
                #cCls.append( mem['code'])
                cCls.append(name)
                cCls.append(mem['objName'])
                cCls.append(int(mem['rechargeCount'])/100.00)
                cCls.append(mem['startDate'])
                cCls.append(mem['endDate'])
                ws.append(cCls)
                count += 1
        wb.save(self.parent_path + '/card.xlsx')





    def write_cardBrand_courseCard_taosubjects(self):
        path = self.parent_path + '/' + self.p_tao 
        with open(path,'r') as f :
            load_dict = json.load(f)

    def write_cardBrand_courseCard_subjects(self):
        path = self.parent_path + '/' + self.p_sub

    def write_card_recharge(self):
        path = self.parent_path + '/' + self.p_rec

    def write_goods_goodsType_listCombineGoods(self):
        path = self.parent_path + '/' + self.p_goods

    def reads(self):
        wb = Workbook()
        ws = wb.active
        wb1 = load_workbook(self.parent_path + '/member.xlsx')
        sheetnames = wb1.get_sheet_names()#获得表单名字
        sheet = wb1.get_sheet_by_name(sheetnames[1])#提取某一名字
        c_d={}
        for x in xrange(2,1000):
            num = sheet.cell(row=x,column=15).value
            name = sheet.cell(row=x,column=1).value
            age = sheet.cell(row=x,column=4).value
            sex = sheet.cell(row=x,column=6).value
            phone = sheet.cell(row=x,column=8).value
            print num
            c_d[str(num)] = [name,age,sex,phone]

        with open(self.parent_path +'/mybqzcard.txt') as f:
            j = json.load(f)
            title = [u'会员号',u'会员名',u'年龄',u'性别',u'电话',u'项目名称',u'类型',u'金额',u'卡总数',u'已用次数',u'剩余次数',u'当前状态',u'购买时间',u'有效时间']
            ws.append(title)
            for i in j:
                my_list = []
                my_list.append(i['memberCode'])
                my_list.append(c_d[str(i['memberCode'])][0])
                my_list.append(c_d[str(i['memberCode'])][1])
                my_list.append(c_d[str(i['memberCode'])][2])
                my_list.append(c_d[str(i['memberCode'])][3])
                my_list.append(i['itemName'])
                my_list.append(i['type'])
                my_list.append(i['amount'])
                my_list.append(i['cardCount'])
                my_list.append(i['useCount'])
                my_list.append(i['remainCount'])
                my_list.append(i['currentStatus'])
                my_list.append(i['buyTime'])
                my_list.append(i['effectiveTime'])
                ws.append(my_list)
        wb.save(self.parent_path+'/user_card.xlsx')
    





if __name__ == '__main__':
    #getExcelSheet()
    phone = '18519005765'
    #phone = '13675846617'
    eh = Exchange(phone)
    eh.action()
    #eh.reads()
    #reads()
