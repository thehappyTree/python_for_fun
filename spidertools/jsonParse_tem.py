# -*- coding:utf-8 -*-
import json
from openpyxl import load_workbook,Workbook

class exchange:
    def __init__(self,parent_path):
        self.parent_path = parent_path

    def reads(self):
        wb = Workbook()
        ws = wb.active
        wb1 = load_workbook(self.parent_path + '/member.xlsx')
        sheetnames = wb1.get_sheet_names()#获得表单名字
        sheet = wb1.get_sheet_by_name(sheetnames[1])#提取某一名字
        c_d={}
        for x in xrange(2,1000):
            num = sheet.cell(row=x,column=15).value
            name = sheet.cell(row=x,column=1).value
            age = sheet.cell(row=x,column=4).value
            sex = sheet.cell(row=x,column=6).value
            phone = sheet.cell(row=x,column=8).value
            print num
            c_d[str(num)] = [name,age,sex,phone]

        with open(self.parent_path +'/banlance.txt') as f:
            j = json.load(f)
            title = [u'会员号',u'会员名',u'年龄',u'性别',u'电话',u'账户总充值',u'充值卡余额',u'当前余额',u'账户总欠款',u'当前欠款',u'积分']
            ws.append(title)
            for i in j:
                my_list = []
                my_list.append(i['memberCode'])
                my_list.append(c_d[str(i['memberCode'])][0])
                my_list.append(c_d[str(i['memberCode'])][1])
                my_list.append(c_d[str(i['memberCode'])][2])
                my_list.append(c_d[str(i['memberCode'])][3])
                my_list.append(i['accountBalance'])
                my_list.append(i['rechargeCardBalance'])
                my_list.append(i['currentBalance'])
                my_list.append(i['accountAmount'])
                my_list.append(i['currentAmount'])
                my_list.append(i['accumulatePoints'])
                ws.append(my_list)
        wb.save(self.parent_path+'/acount_card.xlsx')





if __name__ == '__main__':
    #getExcelSheet()
    #phone = '18519005765'
    phone = '13675846617'
    eh = exchange(phone)
    eh.reads()
    #reads()
