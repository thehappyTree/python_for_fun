# _*_ coding:utf-8 _*_
import sys
import time
import urllib
import urllib2
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
from headers import hds
reload(sys)
sys.setdefaultencoding('utf8')
#some User Agents
#每个nav的信息
#self.dict_mes = {'info':info,'cards':cards,'serviceProduct':serviceProduct,'product':product,'record':record,'balance':balance,'coupon':coupon,'nursing':nursing,'healthy':healthy}
info_list=\
        cards_list=\
        serviceProduct_list=\
        product_list=\
        record_list=\
        balance_list=\
        coupon_list=\
        nursing_list=\
        healthy_list=[]
def meiye_spider():
    try_times = 0
    change_header = 1
    #f = open('lxm','w')

    url = header_url + '/report/customer/list.jhtml'
    #url = 'https://www.douban.com/tag/'+urllib.quote(book_tag)+'?start='+str(page_num*20)+'&type=T'
    time.sleep(np.random.rand()*5)
    plain_text = None


    try:
        print '---connect---'
        req = urllib2.Request(url, headers = hds[change_header%len(hds)])
        source_code = urllib2.urlopen(req).read()
        plain_text = str(source_code)
        #f.write(plain_text)

    except (urllib2.HTTPError, urllib2.URLError), e:
        print e
    soup = BeautifulSoup(plain_text)
    detail = soup.find('tbody')
    trs = detail.find_all('tr')
    for tr in trs:
        k = 0
        for td in tr.find_all('td'):
            if k == 8:
                detail_request = td.find('a').get('href')
                detail_url = header_url + detail_request
                #print td.find('a').href
                req_detail = urllib2.Request(url, headers = hds[change_header%len(hds)])
                print 'connect success'
                source_detail = urllib2.urlopen(req_detail).read()
                detail_text = str(source_detail)
                detail_spider(detail_text)

            else:
                print td.text
            k+=1
        break

    #list_soup = soup.find('div', {'class':'mod book-list'})
    #for d in detail:


    try_times += 1
    #if list_soup == None and try_times < 200:
    #    continue
    #elif list_soup == None or len(list_soup)<=1:
    #    break#200次请求后没有新的内容
    #    try_times=0#set 0 when got valid information
    change_header += 1

#拉取详情信息
def detail_spider(detail_text):
    changer_header += 1
    soup = BeautifulSoup(detail_text)
    ul = soup.find('ul',{'class':'nav-tabs'})
    lis = ul.find_all('li')
    nd = navdeal()
    for li in lis:
        nav_request = li.find('a').get(href)
        nav_url = header_url + nav_request
        req_nav = urllib2.Request(url, headers = hds[change_header%len(hds)])
        source_nav = urllib2.urlopen(req_nav).read()
        nav_text = str(source_nav)
        nav_type = nav_request.split('type=')[1]
        #tmp_dict = dict_mes.get(nav_type)
        # 页面类型
        soup2 = BeautifulSoup(nav_text)
        nd.setText(soup2)
        # 客户详情,当前类型已在navdeal类中定义爬取规则
        if nav_type in nd.dict_mes:
            action = 'nd.' + nav_type +'()'
            eval(action)

        change_header += 1







class navdeal:
    def __init__(self):
        info = \
            cards = \
            serviceProduct=\
            product=\
            record=\
            balance=\
            coupon=\
            nursing=\
            healthy = dict()
        self.dict_mes = {'info':info,'cards':cards,'serviceProduct':serviceProduct,'product':product,'record':record,'balance':balance,'coupon':coupon,'nursing':nursing,'healthy':healthy}

    def setText(bSoup):
        self.bSoup= bSoup

    # 客户详情
    def info(self):
        #self.bSoup.
        div = soup2.find('div',{'class':'panel-body'})
        div_groups = div.find_all('div',{'class':'form-group'})
        for div_group in div_groups:
            k = 1
            for label in div_group.find_all('label'):
                if k % 2 == 0:
                    lable.text()
                k += 1

             






    # 疗程卡
    def cards(self):
        pass

    # 单次
    def serviceProduct(self):
        pass
    
    # 产品
    def product(self):
        pass

    # 账户记录
    def record(self):
        pass

    # 余额/欠款
    def balance(self):
        pass
    
    #优惠券
    def coupon(self):
        pass

    # 护理日志
    def nursing(self):
        pass

    #健康日志
    def healthy(self):
        pass



def print_book_lists_excel(book_lists, book_tag_lists):
    wb = Workbook()
    ws=[]
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i].decode()))

    for i in range(len(book_tag_lists)):
        ws[i].append(['序号','书名','评分', '评价人数','作者','出版社'])
        count=1
        for b1 in book_lists[i]:
            ws[i].append([count, b1[0], float(b1[1]), int(b1[2]),b1[3],b1[4]])
            count+=1
    save_path = 'book_list'

    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i].decode())
    save_path+='.xlsx'
    wb.save(save_path)

if __name__ == '__main__':
    #book_tag_lists = ['科普','经典','生活','心灵','文学']
    #book_tag_lists = ['摄影']
    #book_lists = do_spider(book_tag_lists)
    #print_book_lists_excel(book_lists, book_tag_lists)
    meiye_spider()
        
