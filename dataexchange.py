# _*_ coding:utf-8 _*_
import sys
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook
from pytool import BSoupFactory,navdeal,excelFactory
import os
reload(sys)
sys.setdefaultencoding('utf8')
#some User Agents
#每个nav的信息
#self.dict_mes = {'info':info,'cards':cards,'serviceProduct':serviceProduct,'product':product,'record':record,'balance':balance,'coupon':coupon,'nursing':nursing,'healthy':healthy}
info_list=\
        cards_list=\
        serviceProduct_list=\
        product_list=\
        record_list=\
        balance_list=\
        coupon_list=\
        nursing_list=\
        healthy_list=[]
def meiye_spider():
    #f = open('lxm','w')

    bsf = BSoupFactory()
    soup = bsf.getSoup('index.txt')
    detail = soup.find('tbody')
    trs = detail.find_all('tr')
    excel_header = {'1':'序号','2':'姓名','3':'生日','4':'联系方式'\
            ,'5':'归属门店','6':u'类别','7':'创建时间'}
    contentIndex = []
    for tr in trs:
        line_dict = {}
        for index,td in enumerate(tr.find_all('td')):
            if index == 0 or index == 8:
                continue
            line_dict[str(index)] = td.text
        contentIndex.append(line_dict)
    for c in contentIndex:
        pass

        #print c['2'].replace(' ',''),c['6'].replace(' ','')
        #print c['2'],c['6']


#拉取详情信息
def pull_info():
    # 初始化excel数据
    title = [u'名字',u'体重',u'来源',u'年龄',u'类别',u'性别',u'备注',u'电话',u'生日',u'血型',u'技师',u'门店地址',u'星座',u'体重',u'编号',u'序号',u'门店名']
    eft = excelFactory()
    wb = eft.getExcelObj()
    bsf = BSoupFactory()
    ws = wb.create_sheet(title=u'会员')
    ws.append(title)
    file_name = 'detail_'
    nd = navdeal()
    k=1
    for page in xrange(1,100):
        infosList = []
        for i in xrange(k,1000):
            k = i
            info_file_name = './html/detail/'+file_name + str(page)+ '_' + str(i) + '_info.txt'
            #card_file_name = './html/detail/'+file_name + str(page)+ '_' + str(i) + '_cards.txt'

            if os.path.exists(info_file_name):
                soup_info = bsf.getSoup(info_file_name)
                #print soup
                infoContent = nd.info(soup_info)
                infoContent['id'] = i
                infosList.append(infoContent)
                #print info_file_name
            else:
                break

        #info_sheet = wb.create_sheet
        #ws.append(wb.create_sheet(title=book_tag_lists[i].decode()))

        #ws[i].append(title.values)
        #ws = eft.getExcelSheet(title,wb)
        for c in infosList:
            mes = c.values()
            ws.append(mes)
            print mes
    wb.save('myex.xlsx')



            
#拉取卡详情信息
def pull_cards():
    bsf = BSoupFactory()
    file_name = 'detail_'
    nd = navdeal()
    infosList = []
    k = 1
    for page in xrange(1,100):
        for i in xrange(k,1000):
            #info_file_name = './html/'+file_name + str(page)+ '_' + str(i) + '_info.txt'
            card_file_name = './html/detail/'+file_name + str(page)+ '_' + str(i) + '_cards.txt'
            title = [u'名字',u'体重',u'来源',u'年龄',u'类别',u'性别',u'备注',u'电话',u'生日',u'血型',u'技师',u'门店地址',u'星座',u'体重',u'编号',u'门店名']
            if os.path.exists(card_file_name):
                soup_card = bsf.getSoup(card_file_name)
                #print card_file_name
                #print soup
                cardContent = nd.cards(soup_card)
                for cc in cardContent:
                    for c in cc:
                        print cc[c]
            else:
                k = i

def print_book_lists_excel(content, title):
    pass

    '''
    count=1
    for b1 in book_lists[i]:
        ws[i].append([count, b1[0], float(b1[1]), int(b1[2]),b1[3],b1[4]])
        count+=1
    save_path = 'book_list'

    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i].decode())
    save_path+='.xlsx'
    wb.save(save_path)
    '''



if __name__ == '__main__':
    #book_tag_lists = ['科普','经典','生活','心灵','文学']
    #book_tag_lists = ['摄影']
    #book_lists = do_spider(book_tag_lists)
    #print_book_lists_excel(book_lists, book_tag_lists)
    #meiye_spider()
    method = [pull_cards,pull_info]
    method[1]()
        
