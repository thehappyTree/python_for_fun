# _*_ coding:utf-8 _*_
import sys
from openpyxl import Workbook
from bs4 import BeautifulSoup
import io
import re
reload(sys)
sys.setdefaultencoding('utf8')
#some User Agents
#每个nav的信息
#self.dict_mes = {'info':info,'cards':cards,'serviceProduct':serviceProduct,'product':product,'record':record,'balance':balance,'coupon':coupon,'nursing':nursing,'healthy':healthy}

class navdeal:
    def __init__(self):
        '''
        info = \
            cards = \
            serviceProduct=\
            product=\
            record=\
            balance=\
            coupon=\
            nursing=\
            healthy = {}
        self.dict_mes = {'info':info,'cards':cards,'serviceProduct':serviceProduct,'product':product,'record':record,'balance':balance,'coupon':coupon,'nursing':nursing,'healthy':healthy}
        #info = {'2':}
        #info = {'2':'name','4':'sex','4':'联系方式','5':'归属门店','6':u'类别','7':'创建时间'}
        '''
        self.info_dict = {'11':'name','12':'sex','13':'phone','14':'age','21':'birth','22':'zodiac','23':'btype','24':'height','31':'weight','32':'source','33':'carkind','34':'technician','41':'store','51':'address','61':'note'}
        self.cards_dcit={'1':'item','2':'kind','3':'balance','4':'count','5':'used','6':'left','7':'return','8':'state','9':'buytime','10':'activetime'}


    # 客户详情
    def info(self,soup):
        #self.bSoup.
        div = soup.find('div',{'class':'panel-body'})
        div_groups = div.find_all('div',{'class':'form-group'})
        group = 1
        content = {}
        for div_group in div_groups:
            #print div_group.find_all('label')
            k = 1
            for label in div_group.find_all('label'):
                if k % 2 == 0:
                    ind = k/2
                    index = str(group) + str(ind)
                    title = self.info_dict.get(index,'unkown')
                    text = label.text
                    content[title] = text.replace(' ','')
                k+=1
            group += 1
        #print content
        return content

    # 疗程卡
    def cards(self,soup):
        print soup
        tbody = soup.find('tbody')
        print '----tbody'
        print tbody
        list_content = []
        for tr in tbody.find_all('tr'):
            c_content = {}
            for index, td in enumerate(tr.find_all('td')):  
                title = self.cards_dcit[index+1]
                text = td.text
                c_content[title] = text.replace(' ','')
            list_content.append(c_content)






    # 单次
    def serviceProduct(self):
        pass
    
    # 产品
    def product(self):
        pass

    # 账户记录
    def record(self):
        pass

    # 余额/欠款
    def balance(self):
        pass
    
    #优惠券
    def coupon(self):
        pass

    # 护理日志
    def nursing(self):
        pass

    #健康日志
    def healthy(self):
        pass


class excelFactory:
    def __init__(self):
        pass

    def getExcelObj(self):
        wb = Workbook()
        return wb

    def getExcelSheet(self,title,wb):
        wb = Workbook()
        info_sheet = wb.create_sheet(title=title)
        return ws

class BSoupFactory:
    def __init__(self):
        pass

    def getSoup(self,file_path):
        htmlfile = io.open(file_path, 'r', encoding='utf-8')
        #htmlhandle = htmlfile.read().replace(r'\n|&nbsp|\xa0|\\xa0|\u3000|\\u3000|\\u0020|\u0020','')
        htmlhandle = re.sub(r'\n|&nbsp|\xa0|\\xa0|\u3000|\\u3000|\\u0020|\u0020|" "','',htmlfile.read())
        soup = BeautifulSoup(htmlhandle,features="html.parser")
        return soup

if __name__ == '__main__':
    pass
    #book_tag_lists = ['科普','经典','生活','心灵','文学']
    #book_tag_lists = ['摄影']
    #book_lists = do_spider(book_tag_lists)
    #print_book_lists_excel(book_lists, book_tag_lists)
        
