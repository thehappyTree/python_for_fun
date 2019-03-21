# -*- coding:utf-8 -*-
import csv
from openpyxl import load_workbook, Workbook


class LogTool:
    def __init__(self):
        pass

    def readlog(self):
        wb = load_workbook("withoutContract.xlsx")
        sheet = wb.get_sheet_by_name('Sheet1')

    def user2merchant(self):
        f = open("originuser2merchant.csv", 'r', encoding='UTF-8')
        reader = csv.DictReader(f)
        countMer = {}
        for r in reader:
            print(r['merchantId'])
            userId = r['userId']
            merchantId = r['merchantId']
            if userId in countMer:
                countMer[userId].add(merchantId)
            else:
                tempSet = set()
                tempSet.add(merchantId)
                countMer[userId] = tempSet
        fieldnames  = {'userId','merchantId'}
        f1 = open("user2merchant.csv", "w", newline='')
        f2 = open("user2merchantover1.csv","w", newline='')
        writer1 = csv.DictWriter(f1, fieldnames=fieldnames)
        writer1.writeheader()
        writer2 = csv.DictWriter(f2, fieldnames=fieldnames)
        writer2.writeheader()
        for key, value in countMer.items():
            if len(value) == 1:
                print(userId,value)
                print("----------")

                writer1.writerow({'userId': key, 'merchantId': value.pop()})
            else:

                writer2.writerow({'userId': key, 'merchantId': str(value)})

    def getMerchant(self):
        wb = load_workbook('withoutor.xlsx')
        sheetnames = wb.get_sheet_names()#获得表单名字
        sheet = wb.get_sheet_by_name(sheetnames[0])
        merchantList = []
        for x in range(1, 1050):
            merchant = sheet.cell(row=x, column=1).value
            # 缺少合同的商户id
            merchantList.append(merchant)

        # userid2merchantid
        fu2m = open('user2merchant.csv')
        reader = csv.DictReader(fu2m)
        user2merchant = {}
        for r in reader:
            user2merchant[r['userId']] = r['merchantId']

        merCount = {}
        # 登录日志
        with open("login_log.csv", 'r', encoding='UTF-8') as f:
            reader = csv.DictReader(f)
            for r in reader:
                if r['user_id'] in user2merchant and user2merchant[r['user_id']] in merchantList:
                    merCount[user2merchant[r['user_id']]] = merCount[user2merchant[r['user_id']]] + 1 if user2merchant[r['user_id']] in merCount else 1

        for key, value in merCount.items():
            print(key,value)

    # 一个userid对应的多用户
    def getMerchantMany(self):

        wb = load_workbook('withoutor.xlsx')
        sheetnames = wb.get_sheet_names()#获得表单名字
        sheet = wb.get_sheet_by_name(sheetnames[0])
        merchantList = []
        for x in range(1, 1050):
            merchant = sheet.cell(row=x, column=1).value
            # 缺少合同的商户id
            merchantList.append(merchant)

        # userid2merchantid
        fu2m = open('user2merchantover1.csv')
        reader = csv.DictReader(fu2m)
        user2merchant = {}
        for r in reader:
            user2merchant[r['userId']] = eval(r['merchantId'])

        merCount = {}
        merList = []
        # 登录日志
        with open("login_log.csv", 'r', encoding='UTF-8') as f:
            reader = csv.DictReader(f)
            for r in reader:
                if r['user_id'] in user2merchant:
                        for merchantId in user2merchant[r['user_id']]:
                            if merchantId in merchantList:
                                if  r['user_id'] in merCount:
                                    merCount[r['user_id']][1] = merCount[r['user_id']][1] + 1
                                    continue
                                else:
                                    merCount[r['user_id']] = [user2merchant[r['user_id']],1]
        wb = Workbook()
        ws = wb.active

        k = 1
        for key, value in merCount.items():
            ws.cell(row=k,column=1).value = key
            ws.cell(row=k,column=2).value = ','.join(value[0])
            ws.cell(row=k, column=3).value = value[1]
            print(key, value)
            k += 1
        wb.save('manymerchant.xlsx')





def test():
    f = open('f1.csv', 'w', newline='')
    write = csv.DictReader(f)
    fieldsnames = {'user_id'}
    write = csv.DictWriter(f, fieldnames=fieldsnames)
    write.writeheader()
    write.writerow({'user_id':'2112'})
    write.writerow({'user_id':'2112'})
    write.writerow({'user_id':'2112'})


if __name__ == '__main__':
    #LogTool().user2merchant()
    #LogTool().getMerchant()
    #test()
    LogTool().getMerchantMany()


